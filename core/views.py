from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from rest_framework import mixins, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAdminUser
from .models import *
from .serializers import *

class PublishedViewSet(viewsets.ModelViewSet):
    """Public read endpoints; authenticated staff have full REST CRUD."""
    def get_permissions(self):
        return [AllowAny()] if self.request.method in ('GET', 'HEAD', 'OPTIONS') else [IsAdminUser()]
    def get_queryset(self):
        queryset = self.queryset
        return queryset if self.request.user.is_staff else queryset.filter(is_published=True)

class SiteSettingsViewSet(viewsets.ModelViewSet):
    queryset = SiteSettings.objects.all(); serializer_class = SiteSettingsSerializer
    def get_permissions(self):
        return [AllowAny()] if self.request.method in ('GET', 'HEAD', 'OPTIONS') else [IsAdminUser()]
class HeroSlideViewSet(PublishedViewSet): queryset = HeroSlide.objects.all(); serializer_class = HeroSlideSerializer
class StatisticViewSet(PublishedViewSet): queryset = Statistic.objects.all(); serializer_class = StatisticSerializer
class FacultyViewSet(PublishedViewSet): queryset = Faculty.objects.all(); serializer_class = FacultySerializer
class ArticleViewSet(PublishedViewSet): queryset = Article.objects.all(); serializer_class = ArticleSerializer
class AnnouncementViewSet(PublishedViewSet): queryset = Announcement.objects.all(); serializer_class = AnnouncementSerializer
class InstitutionViewSet(PublishedViewSet): queryset = Institution.objects.all(); serializer_class = InstitutionSerializer
class StudentGroupViewSet(PublishedViewSet): queryset = StudentGroup.objects.all(); serializer_class = StudentGroupSerializer
class CampusBuildingViewSet(PublishedViewSet): queryset = CampusBuilding.objects.all(); serializer_class = CampusBuildingSerializer
class ServiceViewSet(PublishedViewSet): queryset = Service.objects.all(); serializer_class = ServiceSerializer
class GalleryImageViewSet(PublishedViewSet): queryset = GalleryImage.objects.all(); serializer_class = GalleryImageSerializer
class TestimonialViewSet(PublishedViewSet): queryset = Testimonial.objects.all(); serializer_class = TestimonialSerializer
class PageContentViewSet(viewsets.ModelViewSet):
    queryset = PageContent.objects.all(); serializer_class = PageContentSerializer
    def get_permissions(self):
        return [AllowAny()] if self.request.method in ('GET', 'HEAD', 'OPTIONS') else [IsAdminUser()]
class RegistrationViewSet(viewsets.ModelViewSet):
    queryset = Registration.objects.select_related('faculty'); serializer_class = RegistrationSerializer
    def get_permissions(self): return [AllowAny()] if self.action == 'create' else [IsAdminUser()]
    def perform_create(self, serializer):
        serializer.save(status='new')
    def get_queryset(self):
        queryset = super().get_queryset()
        faculty = self.request.query_params.get('faculty')
        status = self.request.query_params.get('status')
        search = self.request.query_params.get('search', '').strip()
        if faculty:
            queryset = queryset.filter(faculty_id=faculty)
        if status:
            queryset = queryset.filter(status=status)
        if search:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(last_name__icontains=search) | Q(first_name__icontains=search) |
                Q(middle_name__icontains=search) | Q(faculty__name__icontains=search)
            )
        return queryset

    @action(detail=False, methods=['get'])
    def export(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Inscriptions'
        headers = ['N°', 'Nom', 'Post-nom', 'Prénom', 'Faculté', 'Motivation', 'Statut', 'Date']
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='075271')
        statuses = dict(Registration._meta.get_field('status').choices)
        for registration in self.get_queryset():
            sheet.append([
                registration.id, registration.last_name, registration.middle_name,
                registration.first_name, registration.faculty.name if registration.faculty else '',
                registration.motivation, statuses.get(registration.status, registration.status),
                registration.created_at.strftime('%d/%m/%Y %H:%M'),
            ])
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(len(str(cell.value or '')) for cell in column) + 2, 48)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="inscriptions-uniluk.xlsx"'
        workbook.save(response)
        return response
